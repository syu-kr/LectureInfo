import json
import openpyxl
from openpyxl.styles import Color, PatternFill

jsonData = {}

with open("convert.json", "r", encoding="utf-8") as f:
  jsonData = json.load(f)

apiJson = {}
apiJson["year"] = "2024"
apiJson["semester"] = "1학기 정규"
apiJson["api"] = sorted(jsonData["data"], key=lambda i: (i["단과대학"], i["학부(과)"], 1 if i["과목명"] != "채플" else -1, int(i["학년"]), int(i["학점"]), i["과목명"], int(i["강좌번호"])))

with open("newConvert.json", "w", encoding="utf-8") as f:
  json.dump(apiJson, f, ensure_ascii=False, indent=2)

excelWB = openpyxl.Workbook()
sheet = excelWB.active

sheet.column_dimensions["A"].width = 9 # 강좌번호
sheet.column_dimensions["B"].width = 9 # 과목코드
sheet.column_dimensions["C"].width = 35 # 과목명
sheet.column_dimensions["D"].width = 18 # 학부(과)
sheet.column_dimensions["E"].width = 5 # 학년
sheet.column_dimensions["F"].width = 9 # 이수구분
sheet.column_dimensions["G"].width = 15 # 영역구분
sheet.column_dimensions["H"].width = 5 # 학점
sheet.column_dimensions["I"].width = 7 # 교수명
sheet.column_dimensions["J"].width = 12 # 수업시간
sheet.column_dimensions["K"].width = 30 # 장소
sheet.column_dimensions["L"].width = 12 # 단과대학
# sheet.column_dimensions["M"].width = 25 # 비고

sheet["A1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["B1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["C1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["D1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["E1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["F1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["G1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["H1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["I1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["J1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["K1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["L1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["M1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["N1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))

sheet.freeze_panes = "A2"

sheet["A1"] = "강좌번호"
sheet["B1"] = "과목코드"
sheet["C1"] = "과목명"
sheet["D1"] = "학부(과)"
sheet["E1"] = "학년"
sheet["F1"] = "이수구분"
sheet["G1"] = "영역구분"
sheet["H1"] = "학점"
sheet["I1"] = "교수명"
sheet["J1"] = "수업시간"
sheet["K1"] = "장소"
sheet["L1"] = "단과대학"
sheet["M1"] = "비고"
sheet["N1"] = "팀티칭여부"

with open("newConvert.json", "r", encoding="utf-8") as f:
  PRES_DATA = json.load(f)
  rowCount = 1
  
  for realData in PRES_DATA["api"]:
    rowCount += 1
    
    sheet["A" + str(rowCount)] = realData["강좌번호"]
    sheet["B" + str(rowCount)] = realData["과목코드"]
    sheet["C" + str(rowCount)] = realData["과목명"]
    sheet["D" + str(rowCount)] = realData["학부(과)"]
    sheet["E" + str(rowCount)] = realData["학년"]
    sheet["F" + str(rowCount)] = realData["이수구분"]
    sheet["G" + str(rowCount)] = realData["영역구분"]
    sheet["H" + str(rowCount)] = realData["학점"]
    sheet["I" + str(rowCount)] = realData["교수명"]
    sheet["J" + str(rowCount)] = realData["수업시간"]
    sheet["K" + str(rowCount)] = realData["장소"]
    sheet["L" + str(rowCount)] = realData["단과대학"]
    sheet["L" + str(rowCount)] = realData["단과대학"]
    sheet["M" + str(rowCount)] = realData["비고"]
    sheet["N" + str(rowCount)] = realData["팀티칭여부"]
    
excelWB.save("newConvert.xlsx")
excelWB.close()